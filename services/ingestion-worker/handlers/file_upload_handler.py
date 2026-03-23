import structlog
from io import BytesIO
from datetime import datetime, timezone
from handlers.base import BaseConnectorHandler
from sentinel_shared.models.media import RawMediaItem
from sentinel_shared.database.session import tenant_context
from sentinel_shared.storage.s3 import S3Client
from sentinel_shared.config import get_settings

logger = structlog.get_logger()

MAX_CONTENT_CHARS = 500_000
MAX_PDF_PAGES = 2000
MAX_EXCEL_ROWS = 500_000


class FileUploadHandler(BaseConnectorHandler):
    """Handler for uploaded PDF/Excel files — downloads from S3, extracts text."""

    async def fetch(
        self,
        config: dict,
        since: str | None,
        location_context: dict | None = None,
    ) -> list[RawMediaItem]:
        files = config.get("files", [])
        if not files:
            logger.warning("file_upload_no_files")
            return []

        settings = get_settings()
        s3 = S3Client()
        tid = tenant_context.get()
        results: list[RawMediaItem] = []

        for file_info in files:
            s3_key = file_info.get("s3_key")
            filename = file_info.get("filename", "unknown")
            content_type = file_info.get("content_type", "")

            if not s3_key:
                logger.warning("file_upload_missing_s3_key", filename=filename)
                continue

            # Validate the S3 key belongs to this tenant (prevent cross-tenant access)
            if not s3_key.startswith(f"{tid}/"):
                logger.error(
                    "file_upload_s3_key_tenant_mismatch",
                    s3_key=s3_key,
                    tenant_id=tid,
                    filename=filename,
                )
                continue

            try:
                logger.info("file_upload_downloading", s3_key=s3_key, filename=filename)
                data = await s3.download_file(settings.s3_uploads_bucket, s3_key)
                logger.info("file_upload_downloaded", s3_key=s3_key, size=len(data))

                text = ""
                metadata: dict = {
                    "filename": filename,
                    "s3_key": s3_key,
                    "content_type": content_type,
                }

                if _is_pdf(filename, content_type):
                    text, page_count = _extract_pdf_text(data)
                    metadata["page_count"] = page_count
                elif _is_xlsx(filename, content_type):
                    text, sheet_count = _extract_xlsx_text(data)
                    metadata["sheet_count"] = sheet_count
                elif _is_xls(filename, content_type):
                    text, sheet_count = _extract_xls_text(data)
                    metadata["sheet_count"] = sheet_count
                else:
                    logger.warning("file_upload_unsupported_type", filename=filename, content_type=content_type)
                    continue

                metadata["char_count"] = len(text)

                # Chunk large files into multiple items
                if len(text) > MAX_CONTENT_CHARS:
                    chunks = _chunk_text(text, MAX_CONTENT_CHARS)
                    logger.info("file_upload_chunking", filename=filename, chunks=len(chunks))
                    for i, chunk in enumerate(chunks):
                        item = RawMediaItem(
                            tenant_id=tid,
                            platform="file_upload",
                            external_id=f"{s3_key}#part{i}",
                            content=chunk,
                            author=filename,
                            published_at=datetime.now(timezone.utc),
                            raw_payload={**metadata, "part": i, "total_parts": len(chunks)},
                        )
                        results.append(item)
                else:
                    item = RawMediaItem(
                        tenant_id=tid,
                        platform="file_upload",
                        external_id=s3_key,
                        content=text,
                        author=filename,
                        published_at=datetime.now(timezone.utc),
                        raw_payload=metadata,
                    )
                    results.append(item)

                logger.info("file_upload_extracted", filename=filename, char_count=len(text))

            except Exception as exc:
                logger.error("file_upload_processing_failed", filename=filename, s3_key=s3_key, error=str(exc))
                continue

        logger.info("file_upload_fetch_complete", total_items=len(results))
        return results


def _is_pdf(filename: str, content_type: str) -> bool:
    return content_type == "application/pdf" or filename.lower().endswith(".pdf")


def _is_xlsx(filename: str, content_type: str) -> bool:
    return (
        content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or filename.lower().endswith(".xlsx")
    )


def _is_xls(filename: str, content_type: str) -> bool:
    return (
        content_type == "application/vnd.ms-excel"
        or filename.lower().endswith(".xls")
    )


def _extract_pdf_text(data: bytes) -> tuple[str, int]:
    """Extract text from PDF bytes using pymupdf. Returns (text, page_count)."""
    import fitz

    doc = fitz.open(stream=data, filetype="pdf")
    if doc.is_encrypted:
        doc.close()
        raise ValueError("PDF is password-protected and cannot be processed")
    total_pages = len(doc)
    if total_pages > MAX_PDF_PAGES:
        doc.close()
        raise ValueError(f"PDF has {total_pages} pages, exceeding the limit of {MAX_PDF_PAGES}")
    pages = []
    for page in doc:
        pages.append(page.get_text())
    page_count = len(pages)
    text = "\n\n".join(pages)
    doc.close()
    return text, page_count


def _extract_xlsx_text(data: bytes) -> tuple[str, int]:
    """Extract text from XLSX bytes using openpyxl. Returns (text, sheet_count)."""
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet_texts = []
    total_rows = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            total_rows += 1
            if total_rows > MAX_EXCEL_ROWS:
                wb.close()
                raise ValueError(f"Excel file exceeds the limit of {MAX_EXCEL_ROWS} total rows")
            cell_values = [str(cell) if cell is not None else "" for cell in row]
            row_text = "\t".join(cell_values)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            sheet_texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    sheet_count = len(wb.sheetnames)
    wb.close()
    text = "\n\n".join(sheet_texts)
    return text, sheet_count


def _extract_xls_text(data: bytes) -> tuple[str, int]:
    """Extract text from XLS (legacy) bytes using xlrd. Returns (text, sheet_count)."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=data)
    sheet_texts = []
    total_rows = 0
    for sheet in wb.sheets():
        rows = []
        for row_idx in range(sheet.nrows):
            total_rows += 1
            if total_rows > MAX_EXCEL_ROWS:
                raise ValueError(f"Excel file exceeds the limit of {MAX_EXCEL_ROWS} total rows")
            cell_values = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
            row_text = "\t".join(cell_values)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            sheet_texts.append(f"[Sheet: {sheet.name}]\n" + "\n".join(rows))
    sheet_count = wb.nsheets
    text = "\n\n".join(sheet_texts)
    return text, sheet_count


def _chunk_text(text: str, max_chars: int) -> list[str]:
    """Split text into chunks of max_chars, breaking at newlines when possible."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + max_chars
        if end >= len(text):
            chunks.append(text[start:])
            break
        # Try to break at a newline within the last 10% of the chunk
        search_start = end - (max_chars // 10)
        newline_pos = text.rfind("\n", search_start, end)
        if newline_pos > start:
            end = newline_pos + 1
        chunks.append(text[start:end])
        start = end
    return chunks
