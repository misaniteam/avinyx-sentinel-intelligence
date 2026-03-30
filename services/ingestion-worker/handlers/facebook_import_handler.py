import uuid
import structlog
from io import BytesIO
from datetime import datetime, timezone
from dateutil import parser as dateparser
from openpyxl import load_workbook
from handlers.base import BaseConnectorHandler
from sentinel_shared.models.media import RawMediaItem
from sentinel_shared.database.session import tenant_context
from sentinel_shared.storage.s3 import S3Client
from sentinel_shared.config import get_settings

logger = structlog.get_logger()

MAX_ROWS = 100_000

EXPECTED_COLUMNS = {
    "title",
    "author",
    "datetime",
    "post_link",
    "reaction_count",
    "comments",
}


def _normalize_header(value: str | None) -> str:
    """Normalize header cell to lowercase stripped string."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_datetime(value) -> datetime:
    """Parse a datetime value flexibly, falling back to now."""
    if value is None:
        return datetime.now(timezone.utc)
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    try:
        dt = dateparser.parse(str(value))
        if dt and dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt or datetime.now(timezone.utc)
    except (ValueError, OverflowError):
        return datetime.now(timezone.utc)


def _safe_str(value) -> str:
    """Convert cell value to string, returning empty string for None."""
    if value is None:
        return ""
    return str(value).strip()


def _safe_int(value) -> int:
    """Convert cell value to int, returning 0 for invalid values."""
    if value is None:
        return 0
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


class FacebookImportHandler(BaseConnectorHandler):
    """Handler for XLSX Facebook post imports — parses rows into RawMediaItems."""

    async def fetch(
        self,
        config: dict,
        since: str | None,
        location_context: dict | None = None,
    ) -> list[RawMediaItem]:
        files = config.get("files", [])
        if not files:
            logger.warning("facebook_import_no_files")
            return []

        settings = get_settings()
        s3 = S3Client()
        tid = tenant_context.get()
        results: list[RawMediaItem] = []

        for file_info in files:
            s3_key = file_info.get("s3_key")
            filename = file_info.get("filename", "unknown")

            if not s3_key:
                logger.warning("facebook_import_missing_s3_key", filename=filename)
                continue

            if not s3_key.startswith(f"{tid}/"):
                logger.error(
                    "facebook_import_s3_key_tenant_mismatch",
                    s3_key=s3_key,
                    tenant_id=tid,
                )
                continue

            try:
                logger.info("facebook_import_downloading", s3_key=s3_key)
                data = await s3.download_file(settings.s3_uploads_bucket, s3_key)

                wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                if ws is None:
                    logger.warning("facebook_import_no_active_sheet", filename=filename)
                    wb.close()
                    continue

                rows_iter = ws.iter_rows(values_only=True)

                # Read and validate header row
                header_row = next(rows_iter, None)
                if header_row is None:
                    logger.warning("facebook_import_empty_file", filename=filename)
                    wb.close()
                    continue

                headers = [_normalize_header(cell) for cell in header_row]
                col_map: dict[str, int] = {}
                for idx, h in enumerate(headers):
                    if h in EXPECTED_COLUMNS:
                        col_map[h] = idx

                missing = EXPECTED_COLUMNS - set(col_map.keys())
                if missing:
                    logger.error(
                        "facebook_import_missing_columns",
                        missing=list(missing),
                        found=list(col_map.keys()),
                        filename=filename,
                    )
                    wb.close()
                    continue

                row_count = 0
                for row in rows_iter:
                    row_count += 1
                    if row_count > MAX_ROWS:
                        logger.warning(
                            "facebook_import_row_limit",
                            filename=filename,
                            limit=MAX_ROWS,
                        )
                        break

                    title = _safe_str(row[col_map["title"]])
                    author = _safe_str(row[col_map["author"]])
                    dt_val = row[col_map["datetime"]]
                    post_link = _safe_str(row[col_map["post_link"]])
                    reaction_count = _safe_int(row[col_map["reaction_count"]])
                    comments = _safe_str(row[col_map["comments"]])

                    # Skip empty rows
                    if not title and not author and not post_link:
                        continue

                    published_at = _parse_datetime(dt_val)

                    # Build content from title and comments
                    content_parts = []
                    if title:
                        content_parts.append(title)
                    if comments:
                        content_parts.append(comments)
                    content = "\n\n".join(content_parts) if content_parts else ""

                    # Use post_link as external_id if available, otherwise generate
                    if post_link:
                        external_id = post_link
                    else:
                        external_id = f"fb_import:{row_count}:{uuid.uuid4()}"

                    item = RawMediaItem(
                        tenant_id=tid,
                        platform="facebook",
                        external_id=external_id,
                        content=content,
                        author=author or None,
                        published_at=published_at,
                        url=post_link or None,
                        engagement={
                            "reactions": reaction_count,
                            "comments_text": comments,
                        },
                        raw_payload={
                            "title": title,
                            "author": author,
                            "datetime": str(dt_val),
                            "post_link": post_link,
                            "reaction_count": reaction_count,
                            "comments": comments,
                            "source": "facebook_import",
                            "row_number": row_count,
                        },
                    )
                    results.append(item)

                wb.close()
                logger.info(
                    "facebook_import_parsed",
                    filename=filename,
                    rows_parsed=row_count,
                    items_created=len(results),
                )

            except Exception as exc:
                logger.error(
                    "facebook_import_processing_failed",
                    filename=filename,
                    s3_key=s3_key,
                    error=str(exc),
                )
                continue

        logger.info("facebook_import_fetch_complete", total_items=len(results))
        return results
